import openpyxl
import json
import os
import sys
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import askyesno

# Sorties JSON — dans app/data/ comme dans le projet original
JSON_SOCIETES = os.path.join("app", "data", "societes.json")
JSON_LOTS     = os.path.join("app", "data", "lots.json")

# Fichier XLSM par défaut
XLSM_DEFAULT  = os.path.join("app", "data", "FICHIER PLANNING FUTURA PAM S072026.xlsm")


def str_safe(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.startswith("=") or s.startswith("#"):
        return ""
    return s


def num_safe(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ".").replace("%", "").strip())
    except (ValueError, TypeError):
        return 0.0


def date_safe(v):
    if v is None:
        return ""
    import datetime
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str_safe(v)


# ── Lecture sociétés depuis feuille Equipe (xlsm) ────────────────────────────

def lire_societes(xlsm_path):
    wb = openpyxl.load_workbook(xlsm_path, read_only=True, keep_vba=True)
    sh = wb["Equipe"]
    societes = {}
    last_nom = None

    for row in sh.iter_rows(min_row=2, values_only=True):
        nom_ste  = str_safe(row[0] if len(row) > 0 else None)
        ce       = str_safe(row[1] if len(row) > 1 else None)
        total_ce = num_safe(row[2] if len(row) > 2 else None)
        ace_nom  = str_safe(row[3] if len(row) > 3 else None)
        nb_pers  = int(num_safe(row[4] if len(row) > 4 else None))
        effectif = int(num_safe(row[6] if len(row) > 6 else None))

        if nom_ste:
            last_nom = nom_ste
            if last_nom not in societes:
                societes[last_nom] = {
                    "nom": last_nom,
                    "ce": ce,
                    "totalHeuresCE": int(total_ce),
                    "aces": []
                }

        if ace_nom and last_nom:
            societes[last_nom]["aces"].append({
                "nom": ace_nom,
                "nbPers": nb_pers,
                "totalHeures": nb_pers * 35,
                "effectifActuel": effectif
            })

    wb.close()
    print(f"  → {len(societes)} sociétés lues depuis Equipe")
    return list(societes.values())


# ── Lecture lots depuis export.XLSX ──────────────────────────────────────────

def lire_lots(export_path):
    wb = openpyxl.load_workbook(export_path, read_only=True)
    sh = wb.active
    lots = []

    for i, row in enumerate(sh.iter_rows(min_row=2, values_only=True)):
        if not row[3]:
            continue
        try:
            cde = int(float(str(row[3])))
        except (ValueError, TypeError):
            continue

        nb_pieces = int(num_safe(row[7]))
        cadence   = num_safe(row[8])

        heures_fichier = num_safe(row[16]) if len(row) > 16 else 0.0
        if heures_fichier > 0:
            heures = round(heures_fichier, 2)
        elif cadence > 0:
            heures = round(nb_pieces / cadence, 2)
        else:
            heures = 0.0

        lots.append({
            "numCDE":        cde,
            "semaine":       str_safe(row[0]),
            "priorite":      int(num_safe(row[2])),
            "typologie":     str_safe(row[4]),
            "affaire":       str_safe(row[5]),
            "valeurVente":   int(num_safe(row[6])),
            "nbPieces":      nb_pieces,
            "cadence":       round(cadence, 2),
            "heures":        heures,
            "statut":        str_safe(row[9]),
            "statutEchant":  str_safe(row[10]),
            "lotACharge":    str_safe(row[11]),
            "estSousDouane": str_safe(row[12]).lower() == "oui",
            "dateReception": date_safe(row[13]),
            "datePaiement":  str_safe(row[14]),
            "commentaire":   str_safe(row[15]),
        })

    wb.close()
    print(f"  → {len(lots)} lots lus depuis export.XLSX")
    return lots


# ── Main ──────────────────────────────────────────────────────────────────────

def export():
    root = Tk()
    root.withdraw()

    # ── Sociétés : si societes.json existe déjà, demander si on le garde ──
    societes = None
    if os.path.exists(JSON_SOCIETES):
        garder = askyesno(
            "societes.json existant",
            f"Le fichier societes.json existe déjà.\n\n"
            f"Voulez-vous le GARDER tel quel ?\n\n"
            f"  OUI  → on garde les sociétés existantes\n"
            f"  NON  → on relit depuis le fichier Excel"
        )
        if garder:
            print(f"  → societes.json conservé (pas de relecture Excel)")
            with open(JSON_SOCIETES, encoding="utf-8") as f:
                societes = json.load(f)
            print(f"  → {len(societes)} sociétés chargées depuis le JSON existant")

    if societes is None:
        # Charger depuis le XLSM
        xlsm_path = XLSM_DEFAULT
        if not os.path.exists(xlsm_path):
            xlsm_path = askopenfilename(
                title="Choisir le fichier XLSM (feuille Equipe — sociétés)",
                filetypes=[("Fichiers Excel", "*.xlsm *.xlsx")]
            )
            if not xlsm_path:
                print("Annulé.")
                sys.exit(0)

        print(f"Lecture sociétés : {xlsm_path}")
        societes = lire_societes(xlsm_path)

    # ── Lots : toujours relire depuis export.XLSX ──────────────────────────
    export_path = askopenfilename(
        title="Choisir le fichier LOTS (export.XLSX)",
        filetypes=[("Fichiers Excel", "*.xlsx *.XLSX")]
    )
    if not export_path:
        print("Annulé.")
        sys.exit(0)

    print(f"Lecture lots : {export_path}")
    lots = lire_lots(export_path)

    # ── Écriture ───────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(JSON_SOCIETES), exist_ok=True)

    with open(JSON_SOCIETES, "w", encoding="utf-8") as f:
        json.dump(societes, f, ensure_ascii=False, indent=2)

    with open(JSON_LOTS, "w", encoding="utf-8") as f:
        json.dump(lots, f, ensure_ascii=False, indent=2)

    print(f"\nExport OK :")
    print(f"  → {JSON_SOCIETES}  ({len(societes)} sociétés)")
    print(f"  → {JSON_LOTS}  ({len(lots)} lots)")


if __name__ == "__main__":
    export()
